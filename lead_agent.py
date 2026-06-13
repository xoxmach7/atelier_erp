"""
Lead Generation Agent — Sheber ERP
Ищет ателье/швейные в 2GIS, генерирует персональные сообщения через Claude API.

Использование:
  python lead_agent.py --key 2GIS_KEY --claude CLAUDE_KEY --city Алматы --count 50
  python lead_agent.py --key 2GIS_KEY --city Алматы --count 50 --no-claude   # без сообщений

Зависимости: pip install requests anthropic openpyxl
"""

import argparse
import json
import time
import sys
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ─── Конфиг ────────────────────────────────────────────────────────────────

CITY_COORDS = {
    "Алматы":  "76.9286,43.2567",
    "Астана":  "71.4704,51.1801",
    "Шымкент": "69.5958,42.3417",
    "Москва":  "37.6176,55.7558",
    "СПб":     "30.3141,59.9386",
}

SEARCH_QUERIES = [
    "ателье штор",
    "шторный салон",
    "пошив штор",
    "жалюзи шторы",
    "карнизы шторы",
]

PITCH_TEMPLATE = """Ты менеджер по продажам SaaS-продукта Sheber ERP — системы управления для шторных ателье.

Продукт решает:
- Хаос в заказах (замеры → КП → производство → монтаж)
- Отсутствие учёта финансов по каждому заказу
- Зависимость от одного дизайнера (вся информация в его голове)

Напиши короткое (3-5 предложений) персональное холодное сообщение для этого бизнеса.

Бизнес: {name}
Адрес: {address}
Категория: {rubric}

Правила:
- Начни с чего-то конкретного про их бизнес, не "Здравствуйте, мы предлагаем..."
- Укажи одну конкретную боль которую решаем
- Призыв к действию: предложи 15-минутный звонок
- Тон: уважительный, не навязчивый, профессиональный
- На русском языке, без смайликов

Напиши только само сообщение, без пояснений."""


# ─── 2GIS API ───────────────────────────────────────────────────────────────

def search_2gis(query: str, location: str, api_key: str, page_size: int = 20, page: int = 1, debug: bool = False) -> list[dict]:
    url = "https://catalog.api.2gis.com/3.0/items"
    # Пробуем разные наборы fields для совместимости с demo-ключом
    params = {
        "q": query,
        "location": location,
        "radius": 100000,
        "key": api_key,
        "fields": "items.point,items.contact_groups,items.contacts,items.url,items.rubrics,items.address,items.name_ex",
        "page_size": page_size,
        "page": page,
        "type": "branch",
    }
    try:
        resp = requests.get(url, params=params, timeout=15)
        if debug and page == 1:
            print(f"\n  [DEBUG] Status: {resp.status_code}")
        
        if resp.status_code != 200:
            print(f"  [2GIS ERROR] {resp.status_code}: {resp.text[:300]}")
            return []
        
        data = resp.json()
        items = data.get("result", {}).get("items", [])
        
        if debug and items:
            print(f"  [DEBUG] Первый объект: {json.dumps(items[0], ensure_ascii=False, indent=2)[:1500]}")
        
        return items
    except Exception as e:
        print(f"  [2GIS ERROR] {query}: {e}")
        return []


def extract_contacts(item: dict) -> dict:
    phone = ""
    website = ""
    instagram = ""

    # Вариант 1: contact_groups (платный ключ)
    for group in item.get("contact_groups", []):
        for contact in group.get("contacts", []):
            ctype = contact.get("type", "")
            value = contact.get("value", "")
            if ctype == "phone" and not phone:
                phone = value
            elif ctype in ("website", "url") and not website:
                if "instagram" in value.lower():
                    instagram = value
                else:
                    website = value
            elif ctype == "instagram" and not instagram:
                instagram = value

    # Вариант 2: contacts (другая структура)
    if not phone:
        for contact in item.get("contacts", []):
            ctype = contact.get("type", "")
            value = contact.get("value", "")
            if ctype == "phone" and not phone:
                phone = value

    # Вариант 3: url напрямую
    if not website and item.get("url"):
        website = item.get("url", "")

    return {"phone": phone, "website": website, "instagram": instagram}


def get_address(item: dict) -> str:
    if item.get("address_name"):
        comment = item.get("address_comment", "")
        return f"{item['address_name']}, {comment}".strip(", ") if comment else item["address_name"]
    return ""


def collect_leads(city: str, api_key: str, max_total: int, debug: bool = False) -> list[dict]:
    location = CITY_COORDS.get(city, CITY_COORDS["Алматы"])
    seen_ids = set()
    leads = []
    per_query = max(max_total // len(SEARCH_QUERIES), 10)

    for i, query in enumerate(SEARCH_QUERIES):
        if len(leads) >= max_total:
            break
        print(f"  Ищем: '{query}'...")
        items = search_2gis(query, location, api_key,
                            page_size=min(per_query, 50),
                            debug=(debug and i == 0))
        for item in items:
            if len(leads) >= max_total:
                break
            uid = item.get("id", "")
            if uid in seen_ids:
                continue
            seen_ids.add(uid)

            contacts = extract_contacts(item)
            address = get_address(item)
            rubric = ""
            rubrics = item.get("rubrics", [])
            if rubrics:
                rubric = rubrics[0].get("name", "") if isinstance(rubrics[0], dict) else str(rubrics[0])

            leads.append({
                "id": uid,
                "name": item.get("name", "Без названия"),
                "address": address,
                "phone": contacts["phone"],
                "website": contacts["website"],
                "instagram": contacts["instagram"],
                "rubric": rubric,
            })
        time.sleep(0.3)

    return leads


# ─── Claude API ─────────────────────────────────────────────────────────────

def generate_message(lead: dict, claude_key: str) -> str:
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=claude_key)
        prompt = PITCH_TEMPLATE.format(
            name=lead["name"],
            address=lead["address"] or "Алматы",
            rubric=lead["rubric"] or "ателье штор",
        )
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        return message.content[0].text.strip()
    except Exception as e:
        return f"[Ошибка: {e}]"


# ─── Excel ──────────────────────────────────────────────────────────────────

def save_excel(leads: list[dict], city: str, output_path: Path, with_messages: bool):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лиды"

    headers = ["№", "Название", "Адрес", "Телефон", "Сайт", "Instagram", "Категория"]
    if with_messages:
        headers.append("Сообщение")

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, lead in enumerate(leads, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=lead["name"])
        ws.cell(row=row, column=3, value=lead["address"])
        ws.cell(row=row, column=4, value=lead["phone"])
        ws.cell(row=row, column=5, value=lead["website"])
        ws.cell(row=row, column=6, value=lead["instagram"])
        ws.cell(row=row, column=7, value=lead["rubric"])
        if with_messages:
            cell = ws.cell(row=row, column=8, value=lead.get("message", ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        if i % 2 == 0:
            bg = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = bg

    widths = [5, 35, 40, 18, 35, 35, 25] + ([70] if with_messages else [])
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\n✅ Сохранено: {output_path}")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--key",      required=True, help="2GIS API key")
    parser.add_argument("--claude",   default="",    help="Anthropic API key")
    parser.add_argument("--city",     default="Алматы")
    parser.add_argument("--count",    type=int, default=30)
    parser.add_argument("--out",      default="")
    parser.add_argument("--no-claude", action="store_true", help="Не генерировать сообщения")
    parser.add_argument("--debug",    action="store_true", help="Показать сырой ответ 2GIS")
    args = parser.parse_args()

    use_claude = bool(args.claude) and not args.no_claude

    if not args.out:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        args.out = f"leads_{args.city}_{ts}.xlsx"

    print(f"\n🔍 Сбор лидов: {args.city}, цель {args.count}")
    if args.debug:
        print("  [DEBUG MODE ON]")
    print("=" * 50)

    leads = collect_leads(args.city, args.key, args.count, debug=args.debug)
    print(f"\n📋 Найдено: {len(leads)}")

    if not leads:
        print("❌ Пусто. Запусти с --debug чтобы увидеть что возвращает 2GIS.")
        sys.exit(1)

    if use_claude:
        print(f"\n✍️  Генерируем сообщения...")
        for i, lead in enumerate(leads, 1):
            print(f"  [{i}/{len(leads)}] {lead['name'][:40]}...")
            lead["message"] = generate_message(lead, args.claude)
            time.sleep(0.1)
    
    save_excel(leads, args.city, Path(args.out), with_messages=use_claude)

    with_phone = sum(1 for l in leads if l["phone"])
    with_insta = sum(1 for l in leads if l["instagram"])
    print(f"\n📊 Итого: {len(leads)} лидов | 📞 {with_phone} с телефоном | 📸 {with_insta} с Instagram")


if __name__ == "__main__":
    main()
